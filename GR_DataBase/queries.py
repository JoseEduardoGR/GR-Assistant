"""
Módulo para ejecutar consultas y generar reportes usando IA
"""

import os
import sys
import yaml
import pandas as pd
from typing import Optional, Union
from pathlib import Path

# Agregar el directorio raíz del proyecto al path
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from assets.engine import OpenRouterEngine
from GR_DataBase.connection import DatabaseConnection


class DatabaseQueries:
    """Ejecuta consultas a la base de datos y genera reportes usando IA."""
    
    def __init__(self, settings_path: str = "settings.yaml", verbose: bool = None):
        self.settings = self._load_settings(settings_path)
        
        # Si verbose no se especifica, leer desde settings.yaml
        if verbose is None:
            verbose = self.settings.get("verbose", True)
        
        self.engine = OpenRouterEngine(settings_path, verbose=verbose)
        self.db = DatabaseConnection()
        self.prompt_path = self.settings.get("prompt_database", "GR_DataBase/context.gr")
        self.verbose = verbose
    
    def _load_settings(self, path: str) -> dict:
        """Carga la configuración desde el archivo YAML."""
        try:
            with open(path) as f:
                return yaml.safe_load(f) or {}
        except FileNotFoundError:
            print(f"[DatabaseQueries] settings.yaml no encontrado en '{path}'")
            return {}
    
    def _load_prompt(self) -> str:
        """Carga el prompt desde el archivo."""
        try:
            # Construir ruta absoluta basada en la ubicación de este archivo
            abs_prompt_path = os.path.join(os.path.dirname(__file__), "context.gr")
            with open(abs_prompt_path) as f:
                return f.read().strip()
        except FileNotFoundError:
            if self.verbose:
                print(f"[DatabaseQueries] Prompt no encontrado en '{abs_prompt_path}'")
            return ""
    
    def connect(self, database: str = None) -> bool:
        """
        Conecta a la base de datos.
        
        Args:
            database: Nombre de la base de datos (opcional)
            
        Returns:
            True si la conexión fue exitosa
        """
        return self.db.connect(database)
    
    def disconnect(self):
        """Desconecta de la base de datos."""
        self.db.disconnect()
    
    def generate_query(self, user_request: str, schema_info: str = None) -> str:
        """
        Genera una consulta SQL basada en la solicitud del usuario usando IA.
        
        Args:
            user_request: Descripción en lenguaje natural de lo que se quiere consultar
            schema_info: Información del esquema de la base de datos (opcional)
            
        Returns:
            Consulta SQL generada
        """
        base_prompt = self._load_prompt()
        
        # Si no hay schema_info, intentar obtenerla
        if not schema_info and self.db.connection:
            schema_info = self._get_schema_info()
        
        full_prompt = f"""{base_prompt}

Esquema de la base de datos:
{schema_info if schema_info else 'No disponible'}

Solicitud del usuario:
{user_request}

Genera SOLO la consulta SQL, sin explicaciones adicionales ni formato markdown.
La consulta debe ser compatible con SQL Server."""
        
        query = self.engine.process(full_prompt).strip()
        
        # Limpiar la consulta (remover markdown si existe)
        query = self._clean_query(query)
        
        # SIEMPRE imprimir la consulta SQL generada para debug
        print(f"[DatabaseQueries] ===== CONSULTA SQL GENERADA =====")
        print(f"{query}")
        print(f"[DatabaseQueries] ===================================")
        
        return query
    
    def _clean_query(self, query: str) -> str:
        """Limpia la consulta removiendo markdown y texto extra."""
        # Remover bloques de código markdown
        if '```' in query:
            lines = query.split('\n')
            cleaned_lines = []
            in_code_block = False
            
            for line in lines:
                if line.strip().startswith('```'):
                    in_code_block = not in_code_block
                    continue
                if in_code_block or not line.strip().startswith('```'):
                    cleaned_lines.append(line)
            
            query = '\n'.join(cleaned_lines)
        
        return query.strip()
    
    def _get_schema_info(self) -> str:
        """Obtiene información del esquema de la base de datos."""
        try:
            tables = self.db.get_tables()
            schema_info = []
            
            for table in tables[:10]:  # Limitar a 10 tablas para no saturar el prompt
                columns = self.db.get_columns(table)
                cols_str = ", ".join([f"{col[0]} ({col[1]})" for col in columns])
                schema_info.append(f"Tabla: {table}\nColumnas: {cols_str}")
            
            return "\n\n".join(schema_info)
        except Exception as e:
            if self.verbose:
                print(f"[DatabaseQueries] Error al obtener schema: {e}")
            return ""
    
    def execute_query(self, query: str) -> pd.DataFrame:
        """
        Ejecuta una consulta SQL y retorna los resultados como DataFrame.
        Lanza excepción si falla, para que la IA la atrape.
        """
        rows, columns = self.db.execute_query(query)
        
        if not rows:
            return pd.DataFrame(columns=columns if columns else [])
            
        df = pd.DataFrame.from_records(rows, columns=columns)
        
        if self.verbose:
            print(f"[DatabaseQueries] Consulta ejecutada: {len(df)} filas obtenidas")
        
        return df
    
    def query_and_report(self, user_request: str, report_type: str = "excel") -> str:
        """
        Genera una consulta (con hasta 5 intentos si falla), la ejecuta y crea un reporte.
        """
        max_attempts = 5
        attempt = 1
        query = ""
        error_history = ""
        df = pd.DataFrame()
        
        schema_info = self._get_schema_info() if self.db.connection else None
        
        while attempt <= max_attempts:
            print(f"[DatabaseQueries] ===== INTENTO {attempt}/{max_attempts} =====")
            print(f"Solicitud: {user_request}")
            
            prompt_context = user_request
            if error_history:
                prompt_context += f"\n\nERROR ANTERIOR:\nLa base de datos devolvió este error al ejecutar tu consulta anterior:\n{error_history}\nPor favor, corrige tu consulta SQL basándote estrictamente en el esquema proporcionado y devuélvela reparada."
                
            query = self.generate_query(prompt_context, schema_info)
            
            try:
                print(f"[DatabaseQueries] Ejecutando: {query}")
                df = self.execute_query(query)
                
                if df.empty:
                    print("[DatabaseQueries] Advertencia: La consulta no retornó datos (tabla vacía o criterios sin coincidencias).")
                else:
                    print("[DatabaseQueries] Consulta exitosa con datos.")
                    
                break # Éxito, salir del bucle
                
            except Exception as e:
                error_history = str(e)
                print(f"[DatabaseQueries] Error de SQL: {error_history}")
                attempt += 1
                
        if attempt > max_attempts:
            raise ValueError(f"No se pudo generar una consulta SQL válida después de {max_attempts} intentos. Último error: {error_history}")
            
        if df.empty:
            raise ValueError("La consulta se ejecutó correctamente pero no retornó resultados. Verifica que existan datos que cumplan los criterios solicitados.")
        
        # Generar reporte según el tipo
        if report_type.lower() == "excel":
            return self._generate_excel_report(df, user_request)
        elif report_type.lower() == "word":
            return self._generate_word_report(df, user_request)
        elif report_type.lower() == "powerpoint":
            return self._generate_powerpoint_report(df, user_request)
        else:
            raise ValueError(f"Tipo de reporte no soportado: {report_type}")
    
    def _generate_excel_report(self, df: pd.DataFrame, description: str) -> str:
        """Genera un reporte Excel usando el módulo xlsx."""
        from GR_Docs.xlsx.excel import ExcelScriptGenerator
        
        # Guardar datos temporalmente en la ruta correcta
        temp_csv = os.path.join(project_root, "GR_Docs", "xlsx", "temp.csv")
        df.to_csv(temp_csv, index=False)
        
        # Generar reporte
        excel_gen = ExcelScriptGenerator(verbose=self.verbose)
        
        # Obtener nombres de columnas reales
        columnas = ", ".join(df.columns.tolist())
        
        prompt = f"""Crea un archivo Excel con los siguientes datos de la base de datos:

{description}

Los datos están en el archivo CSV: {temp_csv}

IMPORTANTE: Las columnas del CSV son exactamente: {columnas}

NO inventes nombres de columnas. Usa SOLO las columnas que existen en el CSV.

Lee el CSV completo sin filtrar y crea un reporte profesional con:
- Lee TODOS los datos del CSV sin filtrar
- Formato de tabla con todas las filas
- Encabezados en negrita
- Colores apropiados
- Ajusta el ancho de las columnas automáticamente

Ejemplo de código:
```python
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

df = pd.read_csv('{temp_csv}')

wb = Workbook()
ws = wb.active
ws.title = 'Reporte'

# Escribir encabezados
for col_num, column_title in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = column_title
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

# Escribir datos
for row_num, row_data in enumerate(df.values, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=cell_value)

# Ajustar ancho de columnas
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

wb.save('ruta_del_archivo.xlsx')
```"""
        
        script_path, xlsx_path = excel_gen.generate_and_execute(prompt)
        
        if self.verbose:
            print(f"[DatabaseQueries] Reporte Excel generado: {xlsx_path}")
        
        return xlsx_path
    
    def _generate_word_report(self, df: pd.DataFrame, description: str) -> str:
        """Genera un reporte Word usando el módulo doc."""
        from GR_Docs.doc.word import WordScriptGenerator
        
        # Convertir DataFrame a texto formateado
        data_text = df.to_string(index=False)
        
        word_gen = WordScriptGenerator(verbose=self.verbose)
        
        prompt = f"""Crea un documento Word con un reporte de base de datos:

Título: {description}

Datos:
{data_text}

Incluye:
- Título principal
- Tabla formateada con los datos
- Resumen de resultados (cantidad de registros)"""
        
        script_path, docx_path = word_gen.generate_and_execute(prompt)
        
        if self.verbose:
            print(f"[DatabaseQueries] Reporte Word generado: {docx_path}")
        
        return docx_path
    
    def _generate_powerpoint_report(self, df: pd.DataFrame, description: str) -> str:
        """Genera un reporte PowerPoint usando el módulo pptx."""
        from GR_Docs.pptx.powerpoint import PowerPointScriptGenerator
        
        # Convertir DataFrame a texto formateado
        data_text = df.head(10).to_string(index=False)  # Primeras 10 filas
        
        pptx_gen = PowerPointScriptGenerator(verbose=self.verbose)
        
        prompt = f"""Crea una presentación PowerPoint con un reporte de base de datos:

Título: {description}

Datos (primeras 10 filas):
{data_text}

Total de registros: {len(df)}

Incluye:
- Diapositiva de título
- Diapositiva con tabla de datos
- Diapositiva con resumen"""
        
        script_path, pptx_path = pptx_gen.generate_and_execute(prompt)
        
        if self.verbose:
            print(f"[DatabaseQueries] Reporte PowerPoint generado: {pptx_path}")
        
        return pptx_path
    
    def __enter__(self):
        """Soporte para context manager."""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Desconecta al salir del context manager."""
        self.disconnect()
